import argparse
import logging
import os
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('errors.txt'),
        logging.StreamHandler()
    ]
)

class SeatingArrangementGenerator:
    def __init__(self, buffer=0, mode='Dense'):
        self.buffer = buffer
        self.mode = mode.upper()
        self.timetable = None
        self.course_rolls = None
        self.room_capacities = None
        self.roll_names = None
        self.room_buildings = None
        
    def load_input_files(self, timetable_path, course_rolls_path, room_capacity_path, roll_names_path=None):
        """Load all input files with data validation"""
        try:
            # Load course_rolls with validation
            self.course_rolls = pd.read_excel(course_rolls_path)
            logging.info(f"Course rolls columns: {', '.join(self.course_rolls.columns)}")
            logging.info(f"Course rolls loaded with {len(self.course_rolls)} entries")
            
            # Validate required columns exist
            if 'Course' not in self.course_rolls.columns or 'RollNo' not in self.course_rolls.columns:
                raise ValueError("course_rolls.xlsx must contain both 'Course' and 'RollNo' columns")
                
            # Convert all RollNo values to strings
            self.course_rolls['RollNo'] = self.course_rolls['RollNo'].astype(str)
            
            # Get unique courses count
            unique_courses = self.course_rolls['Course'].nunique()
            logging.info(f"Found {unique_courses} unique courses in the course rolls file")
            
            # Load other files
            self.timetable = pd.read_excel(timetable_path)
            # Print column names for debugging
            logging.info(f"Timetable columns: {', '.join(self.timetable.columns)}")
            logging.info("Timetable loaded successfully")
            
            self.room_capacities = pd.read_excel(room_capacity_path)
            # Convert Room to string to prevent type issues
            self.room_capacities['Room'] = self.room_capacities['Room'].astype(str)
            logging.info(f"Room capacities columns: {', '.join(self.room_capacities.columns)}")
            logging.info("Room capacities loaded successfully")
            
            if roll_names_path:
                self.roll_names = pd.read_excel(roll_names_path)
                # Display roll_names info for debugging
                logging.info(f"Roll names file columns: {', '.join(self.roll_names.columns)}")
                logging.info(f"Roll names file contains {len(self.roll_names)} entries")
                
                # Sample a few entries to verify the data format
                if not self.roll_names.empty:
                    sample_size = min(5, len(self.roll_names))
                    sample = self.roll_names.head(sample_size)
                    logging.info(f"Sample of roll_names data: \n{sample}")
                
                # Convert RollNo to string
                if 'RollNo' in self.roll_names.columns:
                    self.roll_names['RollNo'] = self.roll_names['RollNo'].astype(str)
                    # Normalize roll numbers - trim whitespace
                    self.roll_names['RollNo'] = self.roll_names['RollNo'].str.strip()
                    
                logging.info("Roll number to name mapping loaded successfully")
            else:
                self.roll_names = None
                logging.info("No roll number to name mapping provided")
                
            self._extract_building_info()
            return True
            
        except Exception as e:
            logging.error(f"Error loading input files: {str(e)}")
            return False
    
    def _extract_building_info(self):
        """Extract building information from room names"""
        self.room_buildings = {}
        for _, row in self.room_capacities.iterrows():
            room = str(row['Room'])  # Ensure room is a string
            # Extract building info based on room number patterns
            if '-' in room:
                building = room.split('-')[0]  # Format "Building-RoomNumber"
            else:
                # Alternate method: first digit(s) in the room number represents the building
                # For rooms like 6101, 8301, etc., we'll use first digit as building
                # This assumes the first character(s) represent the building
                import re
                building_match = re.match(r'^(\d+)', room)
                if building_match:
                    building = f"Building {building_match.group(1)}"
                else:
                    building = "Unknown"
            self.room_buildings[room] = building
    
    def _calculate_effective_capacity(self, capacity):
        """Calculate effective room capacity based on buffer and seating mode"""
        effective = capacity - self.buffer
        if self.mode == 'SPARSE':
            effective = effective // 2
        return max(effective, 1)  # Ensure at least 1 seat
    
    def _get_roll_numbers_for_course(self, course):
        """Get roll numbers for a given course with complete type handling"""
        try:
            # Get all entries for this course
            course_entries = self.course_rolls[self.course_rolls['Course'] == course]
            
            if course_entries.empty:
                logging.warning(f"No roll numbers found for course: {course}")
                return []
            
            # Collect all roll numbers from all entries
            all_rolls = []
            
            for _, entry in course_entries.iterrows():
                roll_value = entry['RollNo']
                
                # Handle case when roll_value is NaN/None
                if pd.isna(roll_value):
                    continue
                    
                # Convert to string first to handle all cases
                roll_str = str(roll_value).strip()
                
                # If the string contains semicolons, split it
                if ';' in roll_str:
                    rolls = [r.strip() for r in roll_str.split(';') if r.strip()]
                    all_rolls.extend(rolls)
                # If it's a single value
                elif roll_str:
                    all_rolls.append(roll_str)
            
            # Log the count for debugging
            logging.info(f"Found {len(all_rolls)} roll numbers for course: {course}")
            
            return all_rolls
                
        except Exception as e:
            logging.error(f"Error processing roll numbers for {course}: {str(e)}")
            return []
    
    def generate_exam_sheet(self, course, room, date, session, roll_names, output_path):
        """Generate exam sheet in the required format"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"{course} Room {room}"
        
        # Header formatting
        header_font = Font(bold=True, size=12)
        border = Border(bottom=Side(border_style="thin"))
        
        # 1. Create header row
        ws.append([f"Course: {course} | Room: {room} | Date: {date} | Session: {session}"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        ws['A1'].font = header_font
        
        # 2. Add column headers
        ws.append(["Roll", "Student Name", "Signature"])
        for col in range(1, 4):
            ws.cell(row=2, column=col).font = Font(bold=True)
            ws.cell(row=2, column=col).border = border
        
        # 3. Add student data
        for roll, name in roll_names.items():
            ws.append([roll, name, ""])  # Empty string for signature
        
        # 4. Add TA and Invigilator sections
        ws.append([])  # Empty row
        for i in range(1, 6):
            ws.append([f"TA{i}", "", ""])
        
        ws.append([])  # Empty row
        for i in range(1, 6):
            ws.append([f"Invigilator{i}", "", ""])
        
        # Adjust column widths
        for col in range(1, 4):  # Only columns A, B, C
            max_length = 0
            column_letter = get_column_letter(col)
            
            # Skip merged cells
            for cell in ws[column_letter]:
                if not isinstance(cell, MergedCell):
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the file
        try:
            # Create directories if they don't exist
            os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
            wb.save(output_path)
        except Exception as e:
            logging.error(f"Error saving file {output_path}: {str(e)}")
    
    def generate_seating_arrangement(self):
        """Main function to generate seating arrangements"""
        try:
            if not all([self.timetable is not None, 
                       self.course_rolls is not None, 
                       self.room_capacities is not None]):
                raise ValueError("Required input files not loaded")
            
            # Validate required columns in timetable
            required_columns = ['Date', 'Session', 'Course']
            missing_columns = [col for col in required_columns if col not in self.timetable.columns]
            if missing_columns:
                missing_cols_str = ', '.join(missing_columns)
                raise ValueError(f"Required columns missing from timetable: {missing_cols_str}")
            
            # Print column names for debugging
            logging.info(f"Timetable columns: {', '.join(self.timetable.columns)}")
            
            # Process timetable - handle possible case differences in column names
            date_col = next((col for col in self.timetable.columns if col.lower() == 'date'), 'Date')
            session_col = next((col for col in self.timetable.columns if col.lower() == 'session'), 'Session')
            
            logging.info(f"Using columns for grouping: {date_col}, {session_col}")
            timetable_groups = self.timetable.groupby([date_col, session_col])
            
            # Create name mapping dictionary if available
            name_mapping = {}
            if self.roll_names is not None:
                # Check for the required columns with flexible naming
                roll_col = None
                name_col = None
                
                # Look for columns that might contain roll numbers and names
                for col in self.roll_names.columns:
                    if col.lower() in ['rollno', 'roll no', 'roll', 'roll_no', 'roll-no', 'student id', 'id']:
                        roll_col = col
                    elif col.lower() in ['name', 'student name', 'studentname', 'full name', 'fullname']:
                        name_col = col
                
                if roll_col and name_col:
                    logging.info(f"Using columns for name mapping: {roll_col} -> {name_col}")
                    # Normalize roll numbers by removing whitespace, etc.
                    roll_numbers = self.roll_names[roll_col].astype(str).str.strip()
                    
                    # Create mapping dictionary
                    name_mapping = dict(zip(roll_numbers, self.roll_names[name_col]))
                    logging.info(f"Created name mapping with {len(name_mapping)} entries")
                    
                    # Sample a few entries for verification
                    sample_items = list(name_mapping.items())[:5]
                    logging.info(f"Sample mapping entries: {sample_items}")
                else:
                    logging.warning(f"Could not find appropriate roll number and name columns. Available columns: {self.roll_names.columns}")
            
            for (date, session), group in timetable_groups:
                logging.info(f"Processing {date} - {session} session")
                
                # Get all courses for this slot
                course_col = next((col for col in group.columns if col.lower() == 'course'), 'Course')
                if course_col not in group.columns:
                    logging.error(f"Course column not found in timetable group. Available columns: {group.columns}")
                    continue
                
                slot_courses = group[course_col].tolist()
                
                # Check for roll number clashes
                self._check_roll_clashes(slot_courses)
                
                # Create room allocation
                slot_allocation, _ = self._allocate_rooms_for_slot(slot_courses)
                
            # Generate output files
                output_dir = "seating_arrangements"
                # Create directory if it doesn't exist
                os.makedirs(output_dir, exist_ok=True)
                
                for alloc in slot_allocation:
                    course = alloc['Course']
                    room = alloc['Room']
                    rolls = alloc['RollNos'].split(';') if isinstance(alloc['RollNos'], str) and alloc['RollNos'] else []
                    
                    # Log the allocation details for debugging
                    logging.info(f"Allocation for {course} in {room} with {len(rolls)} students")
                    
                    # Map roll numbers to names
                    roll_names = {}
                    for roll in rolls:
                        # Normalize roll number by stripping whitespace for matching
                        normalized_roll = roll.strip()
                        # Try to find the name in the mapping
                        if normalized_roll in name_mapping:
                            roll_names[roll] = name_mapping[normalized_roll]
                        else:
                            logging.warning(f"No name found for roll number: {roll}")
                            roll_names[roll] = "Unknown Name"
                    
                    # Log how many names were found/not found
                    found_count = sum(1 for name in roll_names.values() if name != "Unknown Name")
                    logging.info(f"Found names for {found_count} out of {len(rolls)} roll numbers")
                    
                    # Generate output filename
                    try:
                        date_str = pd.to_datetime(date).strftime('%d_%m_%Y')
                    except:
                        # Fallback if date conversion fails
                        date_str = str(date).replace('/', '_').replace('-', '_')
                    
                    session_str = str(session).lower()
                    output_filename = os.path.join(output_dir, f"{date_str}_{course}_{room}_{session_str}.xlsx")
                    
                    # Generate exam sheet
                    self.generate_exam_sheet(
                        course=course,
                        room=room,
                        date=date_str,
                        session=session,
                        roll_names=roll_names,
                        output_path=output_filename
                    )
                    
                    logging.info(f"Generated: {output_filename}")
            
            logging.info("Seating arrangement generation completed successfully")
            return True
            
        except Exception as e:
            logging.error(f"Error generating seating arrangement: {str(e)}")
            return False
    
    def _check_roll_clashes(self, courses):
        """Check if any roll numbers appear in multiple courses"""
        roll_sets = {}
        for course in courses:
            rolls = self._get_roll_numbers_for_course(course)
            roll_sets[course] = set(rolls)
        
        # Compare all pairs of courses
        course_list = list(roll_sets.keys())
        for i in range(len(course_list)):
            for j in range(i+1, len(course_list)):
                course1 = course_list[i]
                course2 = course_list[j]
                intersection = roll_sets[course1] & roll_sets[course2]
                if intersection:
                    logging.warning(f"Roll number clash between {course1} and {course2}: {intersection}")
    
    def _allocate_rooms_for_slot(self, courses):
        """Allocate rooms for courses in a time slot"""
        # Sort courses by student count (descending)
        course_sizes = [(course, len(self._get_roll_numbers_for_course(course))) for course in courses]
        course_sizes.sort(key=lambda x: -x[1])
        
        # Sort rooms by building and capacity
        rooms = []
        for _, row in self.room_capacities.iterrows():
            room = str(row['Room'])  # Ensure room is a string
            capacity = self._calculate_effective_capacity(row['Capacity'])
            building = self.room_buildings[room]
            rooms.append((room, capacity, building))
        rooms.sort(key=lambda x: (x[2], -x[1]))  # Sort by building, then capacity descending
        
        allocations = []
        seats_left_data = []
        
        for course, size in course_sizes:
            rolls = self._get_roll_numbers_for_course(course)
            remaining_students = size
            allocated_rooms = []
            
            # Allocation logic
            room_index = 0
            while remaining_students > 0 and room_index < len(rooms):
                room, capacity, building = rooms[room_index]
                
                if capacity <= 0:
                    room_index += 1
                    continue
                
                allocate = min(remaining_students, capacity)
                rooms[room_index] = (room, capacity - allocate, building)
                remaining_students -= allocate
                allocated_rooms.append((room, allocate))
                
                room_index += 1
            
            if remaining_students > 0:
                logging.warning(f"Not enough capacity for course {course}. {remaining_students} students couldn't be allocated.")
                continue  # Skip this course instead of raising an error
            
            # Prepare allocation records
            roll_index = 0
            for room, count in allocated_rooms:
                if roll_index >= len(rolls):
                    logging.warning(f"Roll index out of bounds for course {course}")
                    break
                    
                room_rolls = rolls[roll_index:roll_index+count]
                roll_index += count
                
                allocations.append({
                    'Course': course,
                    'Room': room,
                    'RollNos': ";".join(room_rolls),
                    'StudentCount': count
                })
        
        # Prepare seats left data
        for room, capacity, building in rooms:
            if capacity > 0:
                seats_left_data.append({
                    'Room': room,
                    'Building': building,
                    'SeatsLeft': capacity
                })
        
        return allocations, seats_left_data

def main():
    parser = argparse.ArgumentParser(description='Generate examination seating arrangements')
    parser.add_argument('--buffer', type=int, default=0, help='Buffer seats to leave in each room')
    parser.add_argument('--mode', type=str, default='Dense', choices=['Dense', 'Sparse'], 
                       help='Seating mode: Dense or Sparse')
    
    args = parser.parse_args()
    
    generator = SeatingArrangementGenerator(buffer=args.buffer, mode=args.mode)
    
    input_files = {
        'timetable_path': 'timetable.xlsx',
        'course_rolls_path': 'course_rolls.xlsx',
        'room_capacity_path': 'room_capacities.xlsx',
        'roll_names_path': 'roll_names.xlsx'
    }
    
    if not generator.load_input_files(**input_files):
        return
    
    if not generator.generate_seating_arrangement():
        return
    
    logging.info("Operation completed successfully!")

if __name__ == '__main__':
    main()